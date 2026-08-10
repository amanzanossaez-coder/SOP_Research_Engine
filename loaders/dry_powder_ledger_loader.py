from pathlib import Path

import openpyxl


# RE-041.4 -- same convention as personal_capacity_facts_loader.py:
# any sheet not in RESERVED_SHEETS is treated as a patrimonio,
# automatically, so a future third patrimonio is added by adding a
# tab, not by touching this file.
RESERVED_SHEETS = {"Notas"}


# Exact "Concepto" text for Section 1's two cells (data/raw/dry_powder_ledger.xlsx),
# matched literally rather than by row position -- a future reordering
# of rows should not silently break this lookup.
EPISODE_START_LABEL = "Fecha inicio episodio (formato Shiller AAAA.MM)"
INITIAL_DRY_POWDER_LABEL = "Pólvora seca inicial del episodio (€)"

SECTION_1_MARKER = "1. EPISODIO ACTUAL"
SECTION_2_MARKER = "2. REGISTRO DE TRAMOS DESPLEGADOS"


def load_dry_powder_ledger_raw(file_path=None):
    """
    RE-041.4 -- raw load of data/raw/dry_powder_ledger.xlsx.

    Returns {patrimonio_name: {"episode_marker": {label: valor},
    "tranches": [{"fecha", "importe", "postura", "nota"}, ...]}}, or
    None if the file is missing. Does not interpret "Pendiente", does
    not compute anything, does not decide whether an episode is
    active -- that is
    engine.dry_powder_ledger_state.build_local_dry_powder_ledger_state()'s
    job, kept separate the same way every other loader in this project
    stays apart from the engine module that interprets it.
    """

    if file_path is None:
        file_path = Path("data/raw/dry_powder_ledger.xlsx")
    else:
        file_path = Path(file_path)

    if not file_path.exists():
        print(f"❌ No existe el fichero: {file_path}")
        return None

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    patrimonios = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name in RESERVED_SHEETS:
            continue

        worksheet = workbook[sheet_name]
        patrimonios[sheet_name] = _read_patrimonio_sheet(worksheet)

    return patrimonios


def _read_patrimonio_sheet(worksheet):

    section1_row = None
    section2_row = None

    for row in worksheet.iter_rows():
        value = row[0].value
        if value is None:
            continue

        text = str(value).strip()

        if text.startswith(SECTION_1_MARKER):
            section1_row = row[0].row
        elif text.startswith(SECTION_2_MARKER):
            section2_row = row[0].row

    episode_marker = {}
    tranches = []

    if section1_row is not None:
        end_row = (
            section2_row - 1
            if section2_row is not None
            else worksheet.max_row
        )

        # +2 skips the section title row itself and its column-header
        # row ("Concepto" / "Valor" / "Fuente" / "Nota").
        for r in range(section1_row + 2, end_row + 1):
            concepto = worksheet.cell(row=r, column=1).value
            if concepto is None:
                continue

            concepto_text = str(concepto).strip()
            if not concepto_text:
                continue

            episode_marker[concepto_text] = worksheet.cell(
                row=r, column=2
            ).value

    if section2_row is not None:

        # +2 skips the section title row and its own column-header row
        # ("Fecha" / "Importe desplegado" / "Postura vigente" / "Nota").
        for r in range(section2_row + 2, worksheet.max_row + 1):
            fecha = worksheet.cell(row=r, column=1).value
            importe = worksheet.cell(row=r, column=2).value
            postura = worksheet.cell(row=r, column=3).value
            nota = worksheet.cell(row=r, column=4).value

            if fecha is None and importe is None and postura is None:
                continue

            tranches.append(
                {
                    "fecha": fecha,
                    "importe": importe,
                    "postura": postura,
                    "nota": nota,
                }
            )

    return {
        "episode_marker": episode_marker,
        "tranches": tranches,
    }
