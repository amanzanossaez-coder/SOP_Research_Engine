from pathlib import Path

import openpyxl


# RE-043.1 -- sheets in this workbook that are not a patrimonio. Any
# other sheet is treated as one, automatically, per Armando's
# requirement that new third-party patrimonios be added by adding a
# tab, not by touching this file.
RESERVED_SHEETS = {"Notas"}


def load_personal_capacity_facts_raw(file_path=None):
    """
    RE-043.1 -- raw load of data/raw/personal_capacity_facts.xlsx.

    Returns a dict {sheet_name: {concepto_text: valor}}, one entry per
    sheet that is not in RESERVED_SHEETS. Does not interpret any value
    as a verifiable fact or a boolean -- that translation is
    engine.personal_capacity_facts_gate.build_local_personal_capacity_facts_inputs()'s
    job, kept separate here the same way loaders/shiller_loader.py
    keeps raw file reading apart from anything engine/drawdown_engine.py
    later does with it.

    sheet_name doubles as the patrimonio identifier -- "AMS", "AML", or
    any future tab name.
    """

    if file_path is None:
        file_path = Path("data/raw/personal_capacity_facts.xlsx")
    else:
        file_path = Path(file_path)

    if not file_path.exists():
        print(f"❌ No existe el fichero: {file_path}")
        return None

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    patrimonios = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name in RESERVED_SHEETS:
            continue

        worksheet = workbook[sheet_name]
        patrimonios[sheet_name] = _read_concepto_valor_map(worksheet)

    return patrimonios


def _read_concepto_valor_map(worksheet):
    """
    Reads column A ("Concepto") / column B ("Valor") pairs across every
    row. Section headers and blank rows produce entries with valor=None
    that nothing ever looks up -- harmless. A repeated Concepto label
    within the same sheet is ambiguous for a text-keyed lookup, so the
    first occurrence wins and the collision is reported rather than
    silently resolved -- this is exactly the failure mode that made two
    identically-labelled "Valoración cualitativa" cells unsafe before
    RE-043.1 renamed them.
    """

    mapping = {}
    duplicates = []

    for row in worksheet.iter_rows():
        if len(row) < 1:
            continue

        concepto = row[0].value
        if concepto is None:
            continue

        concepto_text = str(concepto).strip()
        if not concepto_text:
            continue

        valor = row[1].value if len(row) > 1 else None

        if concepto_text in mapping:
            duplicates.append(concepto_text)
            continue

        mapping[concepto_text] = valor

    if duplicates:
        print(
            f"⚠️  Etiquetas de Concepto repetidas en '{worksheet.title}' "
            "(se usó la primera aparición, se ignoró el resto): "
            f"{sorted(set(duplicates))}"
        )

    return mapping
