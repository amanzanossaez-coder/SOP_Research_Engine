from pathlib import Path

import openpyxl


# RE-032.7 -- same convention as personal_capacity_facts_loader.py /
# dry_powder_ledger_loader.py: any sheet not in RESERVED_SHEETS is a
# patrimonio, automatically.
RESERVED_SHEETS = {"Notas"}

SECTION_MARKER = "REGISTRO DE ATESTACIONES"


def load_human_approval_raw(file_path=None):
    """
    RE-032.7 -- raw load of data/raw/human_approval_attestations.xlsx.
    RE-B (RE-032.10 iteration B) -- extended with column E,
    "autoriza_techo_90".

    Returns {patrimonio_name: [{"fecha", "postura", "crisis_personal",
    "nota", "autoriza_techo_90"}, ...]}, one attestation-event dict per
    non-blank row, or None if the file is missing. Does not interpret
    dates, does not validate postures, does not decide whether
    autoriza_techo_90 actually applies (that depends on the row's own
    postura being Deploy Aggressively, RE-032.10 point 2) -- that
    translation is
    engine.human_approval_state.build_local_human_approval_inputs()'s
    job, kept separate the same way every other loader in this project
    stays apart from the engine module that interprets it.
    """

    if file_path is None:
        file_path = Path("data/raw/human_approval_attestations.xlsx")
    else:
        file_path = Path(file_path)

    if not file_path.exists():
        print(f"❌ No existe el fichero: {file_path}")
        return None

    workbook = openpyxl.load_workbook(file_path, data_only=True)

    patrimonios = {}

    for sheet_name in workbook.sheetnames:
        if sheet_name in RESERVED_SHEETS:
            continue

        worksheet = workbook[sheet_name]
        patrimonios[sheet_name] = _read_attestation_rows(worksheet)

    return patrimonios


def _read_attestation_rows(worksheet):

    section_row = None

    for row in worksheet.iter_rows():
        value = row[0].value
        if value is None:
            continue

        if str(value).strip().startswith(SECTION_MARKER):
            section_row = row[0].row
            break

    if section_row is None:
        return []

    events = []

    # +2 skips the section title row and its own column-header row.
    for r in range(section_row + 2, worksheet.max_row + 1):
        fecha = worksheet.cell(row=r, column=1).value
        postura = worksheet.cell(row=r, column=2).value
        crisis_personal = worksheet.cell(row=r, column=3).value
        nota = worksheet.cell(row=r, column=4).value
        autoriza_techo_90 = worksheet.cell(row=r, column=5).value

        if fecha is None and postura is None:
            continue

        events.append(
            {
                "fecha": fecha,
                "postura": postura,
                "crisis_personal": crisis_personal,
                "nota": nota,
                "autoriza_techo_90": autoriza_techo_90,
            }
        )

    return events
